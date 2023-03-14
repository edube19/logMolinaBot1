import time
from openpyxl import workbook,load_workbook
#ingreso de las credenciales para el LOGIN
def ingreso(page):
    page.get_by_label("Usuario :").click()
    page.get_by_label("Usuario :").fill("jsanchez")
    page.get_by_label("Clave :").click()
    page.get_by_label("Clave :").fill("logmol2")
    print('lleno los datos')
    page.get_by_role("button", name="Ingresar").click()
    print('presiono el boton')

#ingreso a la seccion de ordenes de compra y registro de los datos
def main(page,lista_productos,datos_factura,long_lista_productos):
    elemento_factura = datos_factura[0]
    destino = elemento_factura["destino"]
    opcion = obtener_valor(destino)
    if opcion!= None:
        factura = str(elemento_factura["factura"])
        cod_serie = str(elemento_factura["cod_serie"])
        num_serie = str(elemento_factura["num_serie"])

        page.wait_for_url("http://190.187.186.218/Default.aspx")
        page.get_by_role("link", name="Compras").click()
        page.wait_for_url("http://190.187.186.218/Compras/Default.aspx")
        page.get_by_role("link", name="5. Ordenes de Compra").click()
        page.wait_for_url("http://190.187.186.218/Compras/Ordenes_Com.aspx")

        page.locator("input[name=\"ctl00\\$Main\\$ImageButton1\"]").click()
        page.wait_for_url("http://190.187.186.218/Compras/Orden_Com.aspx")
        page.locator("input[name=\"ctl00\\$Main\\$cod_clipro\"]").click()
        page.locator("input[name=\"ctl00\\$Main\\$cod_clipro\"]").fill(factura)#FACTURA
        print('lleno la factura')
        page.locator("#Main_UpdatePanel1").click()
        page.locator("select[name=\"ctl00\\$Main\\$TIPO_DOC_REF\"]").select_option("FC")
        page.locator("input[name=\"ctl00\\$Main\\$ser_doc_ref\"]").click()
        page.locator("input[name=\"ctl00\\$Main\\$ser_doc_ref\"]").fill(cod_serie)# CODIGO SERIE
        page.locator("input[name=\"ctl00\\$Main\\$num_doc_ref\"]").click()
        page.locator("input[name=\"ctl00\\$Main\\$num_doc_ref\"]").fill(num_serie)#NUMERO DE SERIE
        print('lleno el codigo de factura')
        page.locator("select[name=\"ctl00\\$Main\\$ALMACEN\"]").select_option(str(opcion))
        print('lleno los datos de la cabecera')
        #i=0
        #while i<long_lista_productos:
        for i in range(long_lista_productos):
            elemento_lista = lista_productos[i]
            print(elemento_lista)
            articulo = elemento_lista['producto']
            cantidad = elemento_lista['cantidad']
            precio_unitario = elemento_lista['precio_unitario']
            page.get_by_role("button", name="Adicionar Item").click()
            page.locator("input[name=\"ctl00\\$Main\\$descri_articulo\"]").fill("")
            page.get_by_role("group", name="Orden de Compra").locator("table:has-text(\"Item : Cód.Articulo : Ingresar Codigo Cantidad : Ingresar Cantidad Precio Unitar\")").click()
            page.locator("input[name=\"ctl00\\$Main\\$descri_articulo\"]").click()
            print('ingresando el producto ',articulo)
            page.keyboard.type(articulo, delay=200)
            #page.locator("input[name=\"ctl00\\$Main\\$descri_articulo\"]").fill(articulo_recortado,timeout=30000)
            #page.locator("input[name=\"ctl00\\$Main\\$descri_articulo\"]").click()
            #time.sleep(3)
            page.get_by_text(articulo).click()
            time.sleep(1)
            print('ingresando la cantidad')
            page.locator("input[name=\"ctl00\\$Main\\$cantidad\"]").fill(str(cantidad))
            page.locator("input[name=\"ctl00\\$Main\\$cantidad\"]").press("Tab")
            print('ingresando el precio unitario')
            page.get_by_role("cell", name="Precio Unitario").click()
            page.locator("input[name=\"ctl00\\$Main\\$valor_uni\"]").click()
            time.sleep(1)
            page.locator("input[name=\"ctl00\\$Main\\$valor_uni\"]").fill(str(precio_unitario))
            time.sleep(1)
            page.locator("input[name=\"ctl00\\$Main\\$valor_uni\"]").press("Tab")
            page.get_by_role("button", name="ACEPTAR").click()
            print('se añadio el producto ',articulo)
            time.sleep(1)
            page.locator("input[name=\"ctl00\\$Main\\$descri_articulo\"]").click()
            page.locator('#Main_btnCerrar').click()
            print('Acabo de ingresar los datos del producto ',articulo)
            time.sleep(1)
    else:
        print('Destino incorrecto ...')
    #page.get_by_role("button", name="SALIR").click()
    print('saliendo ... ')

    time.sleep(3)

def nro_filas_columnas(sheet):
    num_rows = 0
    num_cols = 0
    # Itera a través de las filas y columnas de la hoja de Excel
    for row in sheet.iter_rows():
        num_rows += 1
        for cell in row:
            if cell.value:
                num_cols = max(num_cols, cell.column)

    # Imprime el número de filas y columnas que contienen datos
    return num_rows,num_cols

def extraer_datos(filename):
    wb = load_workbook(filename)
    sheet = wb.worksheets[0]#trabajar con la Hoja1
    #nfilas,ncolumnas = nro_filas_columnas(sheet)
    datos = {
        'producto': [],
        'cantidad': [],
        'precio_unitario': [],
    }
    # Itera a través de las filas de la hoja de Excel
    for row in sheet.iter_rows(min_row=4, values_only=True):
        # Agrega los valores de cada columna al diccionario correspondiente
        datos['producto'].append(row[0])
        datos['cantidad'].append(row[1])
        datos['precio_unitario'].append(row[2])
    
    datos_factura = {
        'factura': [],
        'cod_serie': [],
        'num_serie': [],
        'destino': [],
    }
    #extrayendo los datos de la factura
    for row in sheet.iter_rows(min_row=2, max_row=2, values_only=True):
        datos_factura["factura"].append(row[0])
        datos_factura["cod_serie"].append(row[1])
        datos_factura["num_serie"].append(row[2])
        datos_factura["destino"].append(row[3])
    

    # Crea una lista de diccionarios con los valores de todas las columnas
    datos = [dict(zip(datos, valores)) for valores in zip(*datos.values())]#esto es una lista
    datos_factura= [dict(zip(datos_factura, valores)) for valores in zip(*datos_factura.values())]#esto es una lista
    return datos,datos_factura

# Definir una función que devuelva el valor correspondiente a la clave numérica
def obtener_valor(almacen_destino):
    datos = {
    "ALMACEN CENTRAL LA MOLINA": 100,
    "ALMACEN BAR LA MOLINA": 101,
    "ALMACEN COCINA LA MOLINA": 102,
    "EVENTOS LA MOLINA": 103,
    "HOSPEDAJE LA MOLINA": 104,
    "SERVICIOS GENERALES LA MOLINA": 105,
    "TESORERIA MOLINA": 107,
    "LOGISTICA MOLINA": 109,
    "ADMINISTRACION MOLINA": 110
    }
    if almacen_destino in datos:
        return datos[almacen_destino]
    else:
        return None